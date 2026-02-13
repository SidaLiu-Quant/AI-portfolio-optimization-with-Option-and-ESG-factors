import pandas as pd
import numpy as np
from scipy.optimize import minimize
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.model_selection import train_test_split
import warnings
import seaborn as sns
import matplotlib.pyplot as plt
from qmoms import default_params, qmoms_compute
from scipy.stats import gaussian_kde
from scipy.stats import ttest_ind
import time
from sklearn.metrics import mean_squared_error, r2_score
from scipy import stats
from scipy.stats import mstats
from sklearn.utils import resample
from sklearn.decomposition import PCA
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout
from tensorflow.keras.optimizers import Adam
from linearmodels.panel import PanelOLS
from scipy.stats import mstats
from tensorflow.keras.callbacks import EarlyStopping
from PIL import Image, ImageDraw, ImageFont
import os
from statsmodels.tsa.stattools import grangercausalitytests
from tqdm import tqdm
from sklearn.preprocessing import LabelEncoder
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import EarlyStopping
import warnings
warnings.filterwarnings("ignore")
import tensorflow as tf
tf.get_logger().setLevel('ERROR')
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
import copy
import psutil
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import gc

def print_memory_usage():
    process = psutil.Process()
    print(f"taking memory: {process.memory_info().rss / 1024 ** 2:.2f} MB")

def optimize_portfolio_dynamic(returns, initial_weights, risk_free_rate, dates, model_name):
    max_len = 0
    date_pred = dates[12:] 
    for date in date_pred:
        
        current_initial_weights = initial_weights.loc[date]
        
        non_zero_initial_weights = current_initial_weights[current_initial_weights != 0]
        
        max_len = max(max_len, len(non_zero_initial_weights))  
         
        non_zero_returns = returns.loc[date, non_zero_initial_weights.index]
        
        portfolio_return = np.dot(non_zero_initial_weights, non_zero_returns)
        
        print(initial_weights.loc[:date].iloc[:-1])
        Port_return_for_vol = (initial_weights.loc[:date].iloc[:-1]*returns.loc[:date].iloc[:-1]).sum(axis=1)
        portfolio_return_for_vol = pd.DataFrame(Port_return_for_vol)
        portfolio_std_dev = portfolio_return_for_vol.cov().values
        
        sharpe_ratio = (portfolio_return - float(risk_free_rate.loc[date].iloc[0])) / portfolio_std_dev
    
        def objective_function(weights):
            port_return = np.dot(weights, non_zero_returns)
            return -(port_return - float(risk_free_rate.loc[date].iloc[0])) / portfolio_std_dev  
        
        constraints = ({'type': 'eq', 'fun': lambda weights: np.sum(weights) - 1})
        bounds = [(-1, 1)] * len(non_zero_initial_weights)
    
        result = minimize(objective_function, non_zero_initial_weights, method='L-BFGS-B', bounds=bounds, constraints=constraints)
    
        # Normalize the optimized weights
        optimized_weights = result.x 
    
        # Pad the optimized weights with zeros to match the maximum length
        padded_weights = np.zeros(max_len)
        padded_weights[:len(optimized_weights)] = optimized_weights
    
        # Store the optimized and padded weights
        optimized_weights_dict[date] = padded_weights        
        
        max_len2 = max(len(v) for v in optimized_weights_dict.values())
        
        for key, weights in optimized_weights_dict.items():
            if len(weights) < max_len2:
                optimized_weights_dict[key] = np.append(weights, [np.nan] * (max_len2 - len(weights)))
        
        optimized_weights_df = pd.DataFrame(optimized_weights_dict).T
    return optimized_weights_df

def compute_final_weights(window_weights, optimized_weights_list):
    weighted_sum = sum(w * ow for w, ow in zip(window_weights, optimized_weights_list))
    final_weights = weighted_sum / np.sum(window_weights)
    return final_weights

def calculate_alpha(portfolio_returns, market_returns):
    alphas = {}
    for column in portfolio_returns.columns:
        portfolio = portfolio_returns[column]
        X = sm.add_constant(market_returns)
        model = sm.OLS(portfolio, X).fit()
        alpha = model.params['const']
        alphas[column] = alpha
    return alphas

def calculate_significance(portfolio_returns, market_returns):
    p_values = {}
    t_stats = {}
    t_test_results = {}

    for column in portfolio_returns.columns:
        portfolio = portfolio_returns[column]

        # Perform t-test to check if portfolio's excess return is significantly different from 0
        excess_return = portfolio - market_returns
        t_stat, p_value = stats.ttest_1samp(excess_return.dropna(), 0)  
        t_test_results[column] = {'t-stat': t_stat, 'p-value': p_value}

    return t_test_results

def bootstrap_portfolio(features, targets, configurations, n_iterations=1000):
    """
    features: Feature data used for model training and prediction (features_scaled_df).
    targets: The target variables of the model (e.g., portfolio returns or Sharpe ratio).
    configurations: A dictionary containing different model configurations.
    n_iterations: The number of bootstrap iterations.
    """
    bootstrap_results = {}

    for config_name, config in configurations.items():
        model = get_model_instance(config['alpha_estimation_method'])  
        bootstrap_metrics = []

        for i in range(n_iterations):
            X_resample, y_resample = resample(features, targets)

            model.fit(X_resample, y_resample)

            predictions = model.predict(features)

            r2 = r2_score(targets, predictions)
            mse = mean_squared_error(targets, predictions)
            bootstrap_metrics.append((r2, mse))

        r2_mean = np.mean([x[0] for x in bootstrap_metrics])
        mse_mean = np.mean([x[1] for x in bootstrap_metrics])
        bootstrap_results[config_name] = {'r2_mean': r2_mean, 'mse_mean': mse_mean}

    return bootstrap_results

def get_model_instance(model_name, optim_params=None):
    if model_name == 'OLS':
        return LinearRegression()
    elif model_name == 'NeuralNetwork':
        model = Sequential()
        for units in optim_params['hidden_layers']:
            model.add(Dense(units, activation='relu'))
        model.add(Dense(1))  
        model.compile(optimizer=Adam(), loss='mse')
        return model
    else:
        raise ValueError(f"Unsupported model: {model_name}")

def calculate_vif(dataframe):
    vif_data = pd.DataFrame()
    vif_data["feature"] = dataframe.columns
    vif_data["VIF"] = [
        variance_inflation_factor(dataframe.values, i)
        for i in range(dataframe.shape[1])
    ]
    return vif_data

def perform_factor_analysis(portfolio_returns, factors_df):
    factor_analysis_results = {}

    for column in portfolio_returns.columns:
        portfolio = portfolio_returns[column]
        
        X = sm.add_constant(factors_df)
        model = sm.OLS(portfolio, X).fit()

        analysis_table = model.summary2().tables[0]  
        factor_analysis_results[column] = analysis_table

    return factor_analysis_results

def esg_factor_analysis(portfolio_returns, esg_factors):
    esg_analysis_results = {}
    esg_importances = pd.DataFrame()
    for column in portfolio_returns.columns:
        portfolio = portfolio_returns[column]

        # Regress the portfolio returns on ESG factors
        X = sm.add_constant(esg_factors)
        model = sm.OLS(portfolio, X).fit()

        esg_analysis_results[column] = model.summary()
        esg_importances[column] = model.params[1:]
        # calculate the average importance of ESG
    esg_importances['Average Importance'] = esg_importances.mean(axis=1)


    esg_importances_sorted = esg_importances[['Average Importance']].sort_values(by='Average Importance', ascending=False)

    print("ESG Factor Importance across portfolios:")
    print(esg_importances_sorted)

    return esg_importances_sorted

def train_ols_model(X_train, y_train, feature_set_name):
    X_train_const = sm.add_constant(X_train)
    model = sm.OLS(y_train['Return'], X_train_const).fit()
    print(f"\n{feature_set_name} OLS Model Result:")
    print(model.summary())
    
    vif = pd.DataFrame()
    vif["Variable"] = ['const'] + X_train.columns.tolist()
    vif["VIF"] = [variance_inflation_factor(X_train_const.values, i) for i in range(X_train_const.shape[1])]
    print("\n VIF (VIF):")
    print(vif)
    return model

def build_nn_model(hidden_layers, input_shape, activation='swish', dropout_rate=0.2):

    model = Sequential()
    # 首层明确指定input_dim
    model.add(Dense(hidden_layers[0], input_shape=input_shape, activation=activation))
    if dropout_rate:
        model.add(Dropout(dropout_rate))
    for units in hidden_layers[1:]:
        model.add(Dense(units, activation=activation))
        if dropout_rate:
            model.add(Dropout(dropout_rate))

    model.add(Dense(1, activation='linear'))
    model.compile(optimizer='adam', loss='mse')
    return model

def process_features(features_combined, model_type='ols'):
    vif_data = calculate_vif(features_combined)
    high_vif_features = vif_data[vif_data["VIF"] > 10]["feature"].tolist()
    
    if model_type == 'ols':
        return features_combined.drop(high_vif_features, axis=1)
    elif model_type == 'nn':
        scaler = StandardScaler()
        features_scaled = scaler.fit_transform(features_combined)
        return pd.DataFrame(features_scaled, columns=features_combined.columns)
    else:
        raise ValueError("Invalid model type.")

def fit_regression(data, x_col):
    data = data.copy()
    data['const'] = 1
    model = PanelOLS(
        dependent=data['Return'],
        exog=data[[x_col, 'const']],
        entity_effects=False  
    )
    result = model.fit(cov_type='unadjusted')
    slope = result.params[x_col]
    intercept = result.params['const']
    return slope, intercept

def save_table_as_image(df, title, save_path, format_dict=None):
    df_display = df.copy()

    if format_dict:
        for col, fmt in format_dict.items():
            if col in df_display.columns:
                df_display[col] = df_display[col].apply(lambda x: f"{x:{fmt}}" if pd.notnull(x) else "")

    try:
        font = ImageFont.load_default() 
    except IOError:
        print("Error: cannot load default font")
        return

    text = f"{title}\n\n{df_display.to_string(index=False)}"
    lines = text.split('\n')

    max_width = int(max([font.getlength(line) for line in lines]))  
    bbox = font.getbbox("A")  
    line_height = (bbox[3] - bbox[1]) + 4  

    output_dir = os.path.dirname(save_path)
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"create folders: {output_dir}")
        except Exception as e:
            print(f"Error: cannot create categories {output_dir}. {e}")
            return

    img = Image.new('RGB', 
                   (max_width + 40, line_height * len(lines) + 40),
                   color='white')
    draw = ImageDraw.Draw(img)

    y = 20
    for line in lines:
        draw.text((20, y), line, font=font, fill='black')
        y += line_height

    try:
        img.save(save_path)
        print(f"Saved: {save_path}")
    except Exception as e:
        print(f"Error: cannot save pictures to {save_path}. {e}")


def get_feature_matrix(df, group_name):
    if isinstance(df.columns, pd.MultiIndex):
        return df.xs(group_name, axis=1, level=0)
    else:
        return df.filter(regex=f"^{group_name}_")      

def analyze_esg_lag_effects(df_group, group_name, lags, colors, linestyles, save_path):

    os.makedirs(save_path, exist_ok=True)

    results = []
    scatter_data = {}
    reg_lines = {}
    vif_tables = {}

    for lag in lags:
        lag_col = f'ESG_lag{lag}'
        df_group[lag_col] = df_group.groupby(level='Stock')['ESG'].shift(lag)


    for lag in lags:
        lag_col = f'ESG_lag{lag}'
        needed_cols = ['Return', lag_col, 'smfiv']
        
        if lag_col not in df_group.columns:
            print(f"{group_name} missiong columns: {lag_col}")
            continue
            
        df_clean = df_group[needed_cols].dropna()
        if df_clean.empty:
            print(f"{group_name} lag {lag} invalid data，skip")
            continue
            
        scatter_data[lag] = df_clean
        

        try:

            X_vif = df_clean[[lag_col, 'smfiv']] 
            
            vif_values = [
                variance_inflation_factor(X_vif.values, i)
                for i in range(X_vif.shape[1])
            ]
            
            vif_df = pd.DataFrame({
                "Variable": X_vif.columns,
                "VIF": vif_values
            }).sort_values("VIF", ascending=False)
            
            print(f"\n{'='*30} {group_name} Lag {lag} VIF Test {'='*30}")
            print(vif_df.to_string(index=False, float_format=lambda x: f"{x:.2e}"))
            

            vif_tables[lag] = vif_df
            
            plt.figure(figsize=(8, 2), dpi=150)  
            sns.barplot(
                x="VIF", y="Variable", 
                data=vif_df,
                palette="viridis",
                orient="h"
            )
            plt.title(f"Lag {lag} VIF Test - {group_name}", fontsize=10)  
            plt.xlabel("VIF(log)")
            plt.xscale('log')
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.savefig(
                os.path.join(save_path, f"{group_name}_Lag{lag}_ESG-smfiv_VIF.png"), 
                bbox_inches='tight'
            )
            plt.close()
            
        except Exception as e:
            print(f"VIF calculation failed: {str(e)}")
            vif_tables[lag] = None

        if len(df_clean) < 100:
            print(f"{group_name} lag {lag} sample is not enought（n={len(df_clean)}），skip")
            continue

        try:

            y = df_clean['Return']
            X = df_clean[[lag_col]].copy()
            X['const'] = 1  

            model = PanelOLS(
                dependent=y,
                exog=X,
                entity_effects=True  
            )
            result = model.fit(
                cov_type="clustered",  
                cluster_entity=True
            )

            results.append({
                "Lag": lag,
                "Coef": result.params[lag_col],
                "Pval": result.pvalues[lag_col],
                "Tstat": result.tstats[lag_col],
                "R2": result.rsquared,
                "Nobs": result.nobs,
                "StartDate": df_clean.index.get_level_values('Date').min().strftime('%Y-%m'),
                "EndDate": df_clean.index.get_level_values('Date').max().strftime('%Y-%m')
            })

            reg_lines[lag] = (
                result.params[lag_col],  
                result.params['const']   
            )

        except Exception as e:
            print(f"{group_name} lag {lag} regression failed: {str(e)}")
            results.append({
                "Lag": lag,
                "Coef": None,
                "Pval": None,
                "Tstat": None,
                "R2": None,
                "Nobs": len(df_clean),
                "StartDate": None,
                "EndDate": None
            })


    if results:
        df_results = pd.DataFrame(results).set_index('Lag').sort_index()
        df_results = df_results[['Coef', 'Pval', 'R2', 'Nobs', 'StartDate', 'EndDate']]
        

        print(f"\n{'='*30} {group_name} regression summary {'='*30}")
        print(df_results.to_string(
            float_format=lambda x: f"{x:.4f}" if isinstance(x, (int, float)) else str(x)
        ))
        
        df_results.to_csv(
            os.path.join(save_path, f"{group_name}_Regression_Results.csv"),
            float_format="%.4f"
        )
    else:
        df_results = pd.DataFrame()
        print(f"{group_name} invalid regression")


    if reg_lines:
        plt.figure(figsize=(10, 6), dpi=300)
        

        valid_lags = [lag for lag in lags if lag in reg_lines]
        
        for lag in valid_lags:  

            if (lag not in scatter_data) or (f'ESG_lag{lag}' not in scatter_data[lag].columns):
                print(f"{group_name} Lag{lag} Data is not enought for visualization")
                continue
                
            data = scatter_data[lag].reset_index()
            slope, intercept = reg_lines[lag]
            
            try:
                x_min = data[f'ESG_lag{lag}'].min()
                x_max = data[f'ESG_lag{lag}'].max()
                
                if x_min == x_max:
                    x_range = np.array([x_min])
                else:
                    x_range = np.linspace(x_min, x_max, 100)
                    
                y_pred = intercept + slope * x_range
                
                sns.scatterplot(
                    x=f'ESG_lag{lag}', y='Return',
                    data=data,
                    alpha=0.2,
                    color=colors[lag],
                    label=f'Lag={lag} (n={len(data)})', 
                    legend=False
                )
                
                plt.plot(
                    x_range, y_pred,
                    color=colors[lag],
                    linestyle=linestyles[lag],
                    lw=2,
                    label=f'Lag={lag} Fit (β={slope:.3f})'
                )
                
            except Exception as e:
                print(f"{group_name} Lag{lag} visualization failed: {str(e)}")
                continue

        plt.title(f'{group_name} ESG Lag Analysis', fontsize=14, pad=20)
        plt.xlabel('Standardized ESG Scoring（lag）', fontsize=10)
        plt.ylabel('Stocks Returns', fontsize=10)
        plt.grid(True, alpha=0.3)
        plt.legend(
            title='Lags',
            frameon=True,
            facecolor='white',
            bbox_to_anchor=(1.05, 1),
            loc='upper left'
        )
        
        plt.savefig(
            os.path.join(save_path, f"{group_name}_Lagged_Effects.png"),
            bbox_inches='tight'
        )
        plt.close()

    return df_results, scatter_data, reg_lines, vif_tables

def get_window_data(features, dates):

    data_list = []
    for feat in features:

        if feat in scaled_dfs['esg_scaled'].columns:
            df = scaled_dfs['esg_scaled']
        elif feat in scaled_dfs['esg_smfiv_scaled'].columns:
            df = scaled_dfs['esg_smfiv_scaled']

        data_list.append(df.loc[dates, feat])
    return np.column_stack(data_list)

def optimize_portfolio(alphas, cov_matrix, target_return=None):

    n_assets = len(alphas)
    initial_weights = np.ones(n_assets) / n_assets  
   
    constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}] 
    if target_return is not None:
        constraints.append({'type': 'eq', 'fun': lambda w: np.dot(w, alphas) - target_return})

    bounds = [(0, 1) for _ in range(n_assets)]  
   
    def objective(w):
        return w.T @ cov_matrix @ w
   
    result = minimize(
        objective,
        initial_weights,
        method='SLSQP',
        bounds=bounds,
        constraints=constraints
    )
    return result.x

def objective_function(weights, current_returns, current_cov_matrix, risk_free_rate):
    port_return = np.dot(weights, current_returns)
    port_volatility = np.sqrt(np.dot(weights.T, np.dot(current_cov_matrix, weights)))
    
    risk_free_rate_value = risk_free_rate[0]  
    
    sharpe_ratio = (port_return - risk_free_rate_value) / port_volatility
    return -sharpe_ratio

def optimize_portfolio_esg(factors_esg_unique, returns, cov_matrix, risk_free_rate):
    optimized_weights_dict = {}
    
    for date in factors_esg_unique.index:
        current_returns = returns.loc[date].values
        current_cov_matrix = cov_matrix
        
        initial_weights = np.ones(len(current_returns)) / len(current_returns)  
        
        print("current_returns.shape:", current_returns.shape)
        print("current_cov_matrix.shape:", current_cov_matrix.shape)
        print("initial_weights.shape:", initial_weights.shape)
        
        constraints = ({
            'type': 'eq',
            'fun': lambda weights: np.sum(weights) - 1  
        })
        
        bounds = [(0, 1)] * len(current_returns)
        
        result = minimize(objective_function, initial_weights, 
                          args=(current_returns, current_cov_matrix, risk_free_rate),
                          method='SLSQP',
                          bounds=bounds,
                          constraints=constraints)
        
        optimized_weights_dict[date] = result.x
    
    return optimized_weights_dict

def optimize_portfolio_dynamic(returns, initial_weights, risk_free_rate, dates, model_name):
    max_len = 0
    date_pred = dates[12:] 
    optimized_weights_dict = {}
    optimized_weights_df = pd.DataFrame()
    for date in date_pred:
        
        current_initial_weights = initial_weights.loc[date]
        
        non_zero_initial_weights = current_initial_weights[current_initial_weights != 0]
        
        max_len = max(max_len, len(non_zero_initial_weights))  
         
        returns.index = pd.to_datetime(returns.index + "-01")
        non_zero_returns = returns.loc[date, non_zero_initial_weights.index]
        
        portfolio_return = np.dot(non_zero_initial_weights, non_zero_returns)
        
        print(initial_weights.loc[:date].iloc[:-1])
        Port_return_for_vol = (initial_weights.loc[:date].iloc[:-1]*returns.loc[:date].iloc[:-1]).sum(axis=1)
        portfolio_return_for_vol = pd.DataFrame(Port_return_for_vol)
        portfolio_std_dev = portfolio_return_for_vol.cov().values
        
        sharpe_ratio = (portfolio_return - float(risk_free_rate.loc[date].iloc[0])) / portfolio_std_dev
    
        def objective_function(weights):
            port_return = np.dot(weights, non_zero_returns)
            return -(port_return - float(risk_free_rate.loc[date].iloc[0])) / portfolio_std_dev  
        
        constraints = ({'type': 'eq', 'fun': lambda weights: np.sum(weights) - 1})
        bounds = [(-1, 1)] * len(non_zero_initial_weights)
    
        result = minimize(objective_function, non_zero_initial_weights, method='L-BFGS-B', bounds=bounds, constraints=constraints)
    
        # Normalize the optimized weights
        optimized_weights = result.x 
    
        # Pad the optimized weights with zeros to match the maximum length
        padded_weights = np.zeros(max_len)
        padded_weights[:len(optimized_weights)] = optimized_weights
    
        # Store the optimized and padded weights
        optimized_weights_dict[date] = padded_weights        
        optimized_weights_df = pd.DataFrame(optimized_weights_dict).T

        max_len2 = max(len(v) for v in optimized_weights_dict.values())
        
        for key, weights in optimized_weights_dict.items():
            if len(weights) < max_len2:
                optimized_weights_dict[key] = np.append(weights, [np.nan] * (max_len2 - len(weights)))
        

    return optimized_weights_df

def bootstrap_portfolio(features, targets, configurations, n_iterations=1000):

    bootstrap_results = {}

    for config_name, config in configurations.items():

        model = get_model_instance(config['alpha_estimation_method'])  
        bootstrap_metrics = []

        for i in range(n_iterations):
            X_resample, y_resample = resample(features, targets)

            model.fit(X_resample, y_resample)

            predictions = model.predict(features)

            r2 = r2_score(targets, predictions)
            mse = mean_squared_error(targets, predictions)
            bootstrap_metrics.append((r2, mse))

        r2_mean = np.mean([x[0] for x in bootstrap_metrics])
        mse_mean = np.mean([x[1] for x in bootstrap_metrics])
        bootstrap_results[config_name] = {'r2_mean': r2_mean, 'mse_mean': mse_mean}

    return bootstrap_results

def get_model_instance(model_name):
    if model_name == 'RandomForestRegressor':
        return RandomForestRegressor(n_estimators=10, random_state=42)
    elif model_name == 'Lasso':
        return Lasso(alpha=0.0001, max_iter=50000, tol=1e-3)
    elif model_name == 'ElasticNet':
        return ElasticNet(alpha=0.0001, l1_ratio=0.5, max_iter=50000)
    elif model_name == 'GradientBoostingRegressor':
        return GradientBoostingRegressor(n_estimators=10, random_state=42)
    else:
        raise ValueError(f"Unsupported model: {model_name}")
        
class NeuralNetwork:
    def __init__(self, dependent, exog, **model_params):

        self.params = model_params
 
        self.X = tf.convert_to_tensor(exog.values, dtype=tf.float32)
        self.y = tf.convert_to_tensor(dependent.values, dtype=tf.float32)
        
        self.model = self._build_model(model_params)
        
    def _build_model(self, params):
        model = Sequential()
        
        model.add(Dense(
            units=params['network_architecture']['hidden_layers'][0],
            input_shape=(self.X.shape[1],),
            activation=params['network_architecture']['activation']
        ))
        
        for units in params['network_architecture']['hidden_layers'][1:]:
            model.add(Dense(units, activation=params['network_architecture']['activation']))
            model.add(Dropout(params['network_architecture']['dropout_rate']))
        
        model.add(Dense(1))
        
        model.compile(
            optimizer=Adam(learning_rate=params['training_params']['learning_rate']),
            loss=self._custom_loss if params.get('custom_loss') else 'mse'
        )
        
        return model
    
    def _custom_loss(self, y_true, y_pred):
        portfolio_returns = y_true * y_pred
        sharpe = tf.reduce_mean(portfolio_returns) / tf.math.reduce_std(portfolio_returns)
        return -sharpe
    
    def fit(self, **kwargs):
        early_stop = EarlyStopping(patience=5) if self.params['training_params']['early_stopping'] else None
        self.history = self.model.fit(
            self.X, self.y,
            epochs=self.params['training_params']['epochs'],
            batch_size=self.params['training_params']['batch_size'],
            validation_split=0.0,  
            callbacks=[early_stop] if early_stop else None,
            verbose=0
        )
    
    def predict(self, exog): 
        return self.model.predict(exog)


def calculate_cumulative_performance(portfolio_returns):
    cumulative_performance = (1 + portfolio_returns).cumprod()
    return cumulative_performance


def calculate_sharpe_ratio(portfolio_returns, risk_free_rate):
    sharpe_ratio = (np.mean(portfolio_returns) - np.mean(risk_free_rate)/12) / np.std(portfolio_returns)
    return sharpe_ratio

def calculate_max_drawdown(cumulative_performance):
    rolling_max = np.maximum.accumulate(cumulative_performance)
    drawdown    = cumulative_performance / rolling_max - 1
    maxDD = drawdown.min()
    return maxDD
#########
### Step 1: Load the data
all_prices = pd.read_excel(r'/Users/sida/Desktop/ESG/Data/Data.xlsx', sheet_name='Stoxx600')
ivol = pd.read_excel(r'/Users/sida/Desktop/ESG/Data/Data.xlsx', sheet_name='ivol')
ESG = pd.read_excel(r'/Users/sida/Desktop/ESG/Data/Data.xlsx', sheet_name='ESG-Stoxx600')
initial_weights = pd.read_excel(r'/Users/sida/Desktop/ESG/Data/Data.xlsx', sheet_name='Stoxx600_weights')
risk_free_rate = pd.read_excel(r'/Users/sida/Desktop/ESG/Data/Data.xlsx', sheet_name='Risk_Free_Rate')
Stoxx600 = pd.read_excel(r'/Users/sida/Desktop/ESG/Data/Data.xlsx', sheet_name='Stoxx600_index_performance')
### Step 2: Construct option metrics and ESG ratings as factors

# Step 2.1: Set up original portfolio matrix
all_prices = (all_prices
              .set_index(pd.to_datetime(all_prices['Date']).dt.to_period('M').astype(str))
              .drop(columns=['Date'], errors='ignore'))

initial_weights = (initial_weights
                   .set_index(pd.to_datetime(initial_weights['Date']).dt.to_period('M').astype(str))
                   .drop(columns=['Date'], errors='ignore'))


all_prices = all_prices.apply(pd.to_numeric, errors='coerce')

returns = all_prices.pct_change().shift(-1)
returns = returns.fillna(0)
returns = returns.iloc[:-1]

original_portfolio_return = (returns*initial_weights).sum(axis=1)

original_portfolio_return = original_portfolio_return.replace([np.inf, -np.inf], 0)
# Step 2.2: Construct Option Metrics as Macro Factor
ivol.rename(columns={'Unnamed: 0': 'Date'}, inplace=True)
ivol['Date'] = pd.to_datetime(ivol['Date']).dt.strftime('%Y-%m')
ivol.set_index('Date', inplace=True)
implied_vols = ivol.to_numpy()

days = 365
rate = 0.01
grid = np.array([500, 2])
adjusted_params = default_params.copy()
adjusted_params['grid']['number_points'] = 100
mnes = [1.05, 1.10, 1.15, 1.20, 1.25]
start_date = '2016-08-31'
end_date = '2024-06-30'
try:
    results_list = []
    for i in range(implied_vols.shape[0]):
        vol_slice = implied_vols[i, :]
        qout = qmoms_compute(mnes=mnes, vol=vol_slice, days=days, rate=rate, params=adjusted_params)
        results_list.append(qout)

    results_df = pd.DataFrame(results_list)
    results_df.index = pd.date_range(start=start_date, end=end_date, freq='ME').strftime('%Y-%m') 

except Exception as e:
    print(f"An error occurred: {e}")
    import traceback
    traceback.print_exc()

factors_name = ['smfiv', 'mfis', 'tlm_delta20']
results_df.columns
factors_op = results_df[factors_name]

# Step 2.3: Construct ESG factors as individual stock factor
ESG['Date'] = pd.to_datetime(ESG['Date'])
ESG = ESG.sort_values('Date')

for column in ESG.columns:
    if column != 'Date':  
        ESG[column] = ESG[column].replace(0, np.nan).ffill()
        
esg_monthly = (
    ESG
    .groupby(pd.Grouper(key='Date', freq='M'))  
    .last()                                     
    .reset_index()                             
)

factors_esg = (
    esg_monthly.assign(Date=pd.to_datetime(esg_monthly['Date']).dt.strftime('%Y-%m'))  
    .set_index('Date')                                                 
    .apply(pd.to_numeric, errors='coerce')                             
    .replace(0, np.nan)                                                
    .ffill()                                                                                               
    .replace([np.inf, -np.inf], np.nan)                                
    .fillna(0)                                                         
)

# Step 2.4 Filter Out High ESG sensitive stocks
regression_results = pd.DataFrame(columns=['Stock', 'R_squared', 'Coefficient', 'P_value'])
results_list = []
for stock in returns.columns:
    if stock not in factors_esg.columns:
        print(f"Warning!：Stocks {stock} lack ESG data, already skipped")
        continue
    
    combined_data = pd.DataFrame({
        'ESG': factors_esg[stock],
        'Return': returns[stock]
    }).dropna()
    
    if len(combined_data) < 2:
        print(f"Warning!：Stocks {stock} lack effective data{len(combined_data)}，already skipped")
        continue

    X = sm.add_constant(combined_data['ESG'])  
    y = combined_data['Return']                
    model = sm.OLS(y, X)
    results = model.fit()
    
    results_list.append({
        'Stock': stock,
        'R_squared': results.rsquared,
        'Coefficient': results.params['ESG'],
        'P_value': results.pvalues['ESG']
    })

regression_results = pd.DataFrame(results_list)

significant_stocks = regression_results[
    (regression_results['P_value'] < 0.05) & 
    (regression_results['R_squared'] > 0.025)
].sort_values(by='R_squared', ascending=False)


print("List of High Sensitive group:")
display(significant_stocks)

median_coeff = significant_stocks['Coefficient'].median()

high_group = significant_stocks[significant_stocks['Coefficient'] > median_coeff]
low_group = significant_stocks[significant_stocks['Coefficient'] <= median_coeff]

print(f"Number of High Score group: {len(high_group)}, Number of Low Score group: {len(low_group)}")

plt.figure(figsize=(8,6))
significant_stocks['Coefficient'].hist(bins=20)
plt.xlabel("ESG Coefficient")
plt.ylabel("Count")
plt.title("Distribution of high ESG coefficients")
plt.grid(True)

plt.savefig("/Users/sida/Desktop/ESG/Data/esg_coefficient_count.png", dpi=300, bbox_inches='tight')

plt.show()

# Step 3: Build Models for high ESG sensitive stocks
# 3.1 Prepare data set for variables
y_long = returns.stack().reset_index()
y_long.columns = ['Date', 'Stock', 'Return']

significant_stocks_list = significant_stocks['Stock'].tolist()

y_long_significant = y_long[y_long['Stock'].isin(significant_stocks_list)]

y_long_clean = y_long.replace([np.inf, -np.inf], np.nan).dropna(subset=['Return'])

y_long_significant_clean = y_long_significant.replace([np.inf, -np.inf], np.nan).dropna(subset=['Return'])

y_long_clean['Return'] = mstats.winsorize(
    y_long_clean['Return'], limits=[0.01, 0.01]
)

y_long_significant_clean['Return'] = mstats.winsorize(
    y_long_significant_clean['Return'], limits=[0.01, 0.01]
)

print(f"winsorized Total Sample Average Return: {y_long_clean['Return'].mean():.4f} ± {y_long_clean['Return'].std():.4f}")
print(f"winsorized High Sensitive ESG Sample Average Return: {y_long_significant_clean['Return'].mean():.4f} ± {y_long_significant_clean['Return'].std():.4f}")

# for ESG only
# for OLS model 
factors_esg_long = factors_esg.stack().reset_index()
factors_esg_long.columns = ['Date', 'Stock', 'ESG']

merged_data_esg = pd.merge(
    factors_esg_long[['Date', 'Stock', 'ESG']],
    y_long[['Date', 'Stock', 'Return']],
    on=['Date', 'Stock'],
    how='inner' 
)

# 3.1.1 Prepare Data for significant stock (High Sensitive ESG)
factors_esg_significant = factors_esg[significant_stocks_list]

factors_esg_long_significant = factors_esg_significant.stack().reset_index()
factors_esg_long_significant.columns = ['Date', 'Stock', 'ESG']

merged_data_esg_significant = pd.merge(
    factors_esg_long_significant[['Date', 'Stock', 'ESG']],
    y_long_significant[['Date', 'Stock', 'Return']],  
    on=['Date', 'Stock'],
    how='inner'
)

filtered_data_esg = merged_data_esg_significant[merged_data_esg_significant['ESG'] != 0]
filtered_data_esg = filtered_data_esg.sort_values(['Date', 'Stock'])

filtered_data_esg["Date"] = pd.to_datetime(
    filtered_data_esg["Date"], 
    format="%Y-%m"  
)

filtered_data_esg['Return'] = filtered_data_esg['Return'].replace([np.inf, -np.inf], 0)

filtered_data_esg['Return'] = mstats.winsorize(
    filtered_data_esg['Return'], limits=[0.01, 0.01]
)

mean_return = y_long_clean.groupby('Stock')['Return'].mean().reset_index()
mean_return.columns = ['Stock', 'Avg_Return']

mean_ESG = factors_esg_long_significant.groupby('Stock')['ESG'].mean().reset_index()
mean_ESG.columns = ['Stock', 'Avg_ESG']

merged = pd.merge(mean_ESG, mean_return, on='Stock', how='inner')

plt.figure(figsize=(10, 8))
plt.scatter(merged['Avg_ESG'], merged['Avg_Return'], alpha=0.7)

for i in range(len(merged)):
    plt.text(merged['Avg_ESG'][i], merged['Avg_Return'][i], merged['Stock'][i],
             fontsize=8, ha='right', va='bottom')

plt.xlabel("Avg_ESG")
plt.ylabel("Expected Return")
plt.title("Average ESG vs Expected Return")
plt.grid(True)

plt.savefig("/Users/sida/Desktop/ESG/Data/esg_mean_and_expected_returns.png", dpi=300, bbox_inches='tight')

plt.show()

# Conclusion 2: High Score High Sensitive ESG Sample
latest_esg = factors_esg_long.groupby('Stock')['ESG'].last().reset_index()
significant_stocks = pd.merge(significant_stocks, latest_esg[['Stock', 'ESG']], on='Stock', how='left')

median_esg = significant_stocks['ESG'].median()
high_esg_group = significant_stocks[significant_stocks['ESG'] > median_esg]
low_esg_group = significant_stocks[significant_stocks['ESG'] <= median_esg]
print(f"Number of High ESG group: {len(high_esg_group)}, Number of low ESG: {len(low_esg_group)}")

merged_high_esg = pd.merge(high_esg_group[['Stock']], y_long_significant, on='Stock', how='inner')
merged_low_esg = pd.merge(low_esg_group[['Stock']], y_long_significant, on='Stock', how='inner')

mean_high = merged_high_esg['Return'].mean()
std_high = merged_high_esg['Return'].std()
mean_low = merged_low_esg['Return'].mean()
std_low = merged_low_esg['Return'].std()
print(f"High ESG group return: {mean_high:.4f} ± {std_high:.4f}")
print(f"Low ESG group return: {mean_low:.4f} ± {std_low:.4f}")

# 3.1.2 Prepare data for significant stock (ESG high score)
high_esg_group_list = high_esg_group['Stock'].tolist()
high_esg_group_significant = factors_esg[high_esg_group_list]

y_high_esg_significant = y_long[y_long['Stock'].isin(high_esg_group_list)]

high_esg_group_long_significant = high_esg_group_significant.stack().reset_index()
high_esg_group_long_significant.columns = ['Date', 'Stock', 'ESG']

merged_data_high_esg_significant = pd.merge(
    high_esg_group_long_significant[['Date', 'Stock', 'ESG']],
    y_high_esg_significant[['Date', 'Stock', 'Return']],  
    on=['Date', 'Stock'],
    how='inner'
)

filtered_data_high_esg = merged_data_high_esg_significant[merged_data_high_esg_significant['ESG'] != 0]
filtered_data_high_esg = filtered_data_high_esg.sort_values(['Date', 'Stock'])

filtered_data_high_esg["Date"] = pd.to_datetime(
    filtered_data_high_esg["Date"], 
    format="%Y-%m"  
)

filtered_data_high_esg['Return'] = filtered_data_high_esg['Return'].replace([np.inf, -np.inf], 0)

# 3.1.3 Prepare data for significant stock (ESG low score) 
low_esg_group_list = low_esg_group['Stock'].tolist()
low_esg_group_significant = factors_esg[low_esg_group_list]

y_low_esg_significant = y_long[y_long['Stock'].isin(low_esg_group_list)]

low_esg_group_long_significant = low_esg_group_significant.stack().reset_index()
low_esg_group_long_significant.columns = ['Date', 'Stock', 'ESG']

merged_data_low_esg_significant = pd.merge(
    low_esg_group_long_significant[['Date', 'Stock', 'ESG']],
    y_low_esg_significant[['Date', 'Stock', 'Return']],  
    on=['Date', 'Stock'],
    how='inner'
)

filtered_data_low_esg = merged_data_low_esg_significant[merged_data_low_esg_significant['ESG'] != 0]
filtered_data_low_esg = filtered_data_low_esg.sort_values(['Date', 'Stock'])

filtered_data_low_esg["Date"] = pd.to_datetime(
    filtered_data_low_esg["Date"], 
    format="%Y-%m"  
)

filtered_data_low_esg['Return'] = filtered_data_low_esg['Return'].replace([np.inf, -np.inf], 0)


# for ESG + Option
factors_op_reset = factors_op.reset_index().rename(columns={'index': 'Date'})
## 3.2.0.Prepare Data for High Sensitive ESG + Option
merged_double_step1 = pd.merge(
    factors_esg_long,
    factors_op_reset,
    on=['Date'],  
    how='left'
)

merged_data_double = pd.merge(
    merged_double_step1,
    y_long[['Date', 'Stock', 'Return']],
    on=['Date', 'Stock'],  
    how='left'
)

merged_data_double['Date'] = pd.to_datetime(merged_data_double['Date'] + '-01', format='%Y-%m-%d')
merged_data_double['Date'] = merged_data_double['Date'].dt.strftime('%Y-%m')

filtered_data_double = merged_data_double[merged_data_double['ESG'] != 0]
filtered_data_double = filtered_data_double.sort_values(['Date', 'Stock'])
filtered_data_double = filtered_data_double.set_index(['Stock', 'Date'])

## 3.2.1. High Sensitive ESG + Option
merged_data_double_significant_step1 = pd.merge(
    factors_esg_long_significant,
    factors_op_reset,
    on='Date',
    how='left'
)

merged_data_double_significant = pd.merge(
    merged_data_double_significant_step1,
    y_long[['Date', 'Stock', 'Return']],
    on=['Date', 'Stock'],  
    how='left'
)
merged_data_double_significant['Date'] = pd.to_datetime(merged_data_double_significant['Date'] + '-01', format='%Y-%m-%d')
merged_data_double_significant['Date'] = merged_data_double_significant['Date'].dt.strftime('%Y-%m')

filtered_data_double_significant = merged_data_double_significant[merged_data_double_significant['ESG'] != 0]
filtered_data_double_significant = filtered_data_double_significant.sort_values(['Date', 'Stock'])

## 3.2.2. High Sensitive High Score ESG + Option
merged_data_double_high_significant_step1 = pd.merge(
    high_esg_group_long_significant,
    factors_op_reset,
    on='Date',
    how='left'
)

merged_data_double_high_significant = pd.merge(
    merged_data_double_high_significant_step1,
    y_long[['Date', 'Stock', 'Return']],
    on=['Date', 'Stock'], 
    how='left'
)



filtered_data_double_high_significant = merged_data_double_high_significant[merged_data_double_high_significant['ESG'] != 0]
filtered_data_double_high_significant = filtered_data_double_high_significant.sort_values(['Date', 'Stock'])

## 3.2.3.High Sensitive Low Score ESG + Option
merged_data_double_low_significant_step1 = pd.merge(
    merged_data_low_esg_significant,
    factors_op_reset,
    on='Date',
    how='left'
)

merged_data_double_low_significant = pd.merge(
    merged_data_double_low_significant_step1,
    y_long[['Date', 'Stock', 'Return']],
    on=['Date', 'Stock'], 
    how='left'
)


filtered_data_double_low_significant = merged_data_double_low_significant[merged_data_double_low_significant['ESG'] != 0]
filtered_data_double_low_significant = filtered_data_double_low_significant.sort_values(['Date', 'Stock'])

# 3.3 Split the data
# for ESG only
# for OLS
# 3.3.0 All ESG 

# 3.3.1 High Sensitive ESG
split_date = filtered_data_esg['Date'].unique()[int(len(filtered_data_esg['Date'].unique()) * 0.8)]

train_data_esg = filtered_data_esg[filtered_data_esg["Date"] <= split_date]
test_data_esg = filtered_data_esg[filtered_data_esg["Date"] > split_date]

train_data_esg = train_data_esg.set_index(["Stock", "Date"]).sort_index()
test_data_esg = test_data_esg.set_index(["Stock", "Date"]).sort_index()

# 3.2.2 High Sensitive High Score ESG
split_date_high = filtered_data_high_esg['Date'].unique()[int(len(filtered_data_high_esg['Date'].unique()) * 0.8)]

train_data_high_esg = filtered_data_high_esg[filtered_data_high_esg["Date"] <= split_date_high]
test_data_high_esg = filtered_data_high_esg[filtered_data_high_esg["Date"] > split_date_high]

train_data_high_esg = train_data_high_esg.set_index(["Stock", "Date"]).sort_index()
test_data_high_esg = test_data_high_esg.set_index(["Stock", "Date"]).sort_index()

# 3.2.3 High Sensitive Low Score ESG
split_date_low = filtered_data_low_esg['Date'].unique()[int(len(filtered_data_low_esg['Date'].unique()) * 0.8)]

train_data_low_esg = filtered_data_low_esg[filtered_data_low_esg["Date"] <= split_date_low]
test_data_low_esg = filtered_data_low_esg[filtered_data_low_esg["Date"] > split_date_low]

train_data_low_esg = train_data_low_esg.set_index(["Stock", "Date"]).sort_index()
test_data_low_esg = test_data_low_esg.set_index(["Stock", "Date"]).sort_index()

# for ESG + Option
# for OLS
# 0. All ESG
split_date_double = filtered_data_double.index.get_level_values('Date').unique()[int(len(filtered_data_double.index.get_level_values('Date').unique()) * 0.8)]
train_data_double = filtered_data_double[filtered_data_double.index.get_level_values('Date') <= split_date_double]
test_data_double = filtered_data_double[filtered_data_double.index.get_level_values('Date') > split_date_double]

X_ESG_Op_train = train_data_double[['ESG', 'smfiv', 'mfis', 'tlm_delta20', 'Return']]
X_ESG_Op_test = test_data_double[['ESG', 'smfiv', 'mfis', 'tlm_delta20', 'Return']]
X_ESG_Op_train_clean = X_ESG_Op_train.replace([np.inf, -np.inf], np.nan).dropna()

# 1. High Sensitive ESG+op
dates = merged_data_double_significant['Date']
if not pd.api.types.is_datetime64_any_dtype(dates):
    dates = pd.to_datetime(dates)

merged_data_double_significant['Date'] = pd.to_datetime(
    merged_data_double_significant['Date'], 
    format='%Y-%m'  
)

unique_dates = np.sort(merged_data_double_significant['Date'].unique())

split_idx = int(len(unique_dates) * 0.8)
split_date = unique_dates[split_idx]

X_ESG_Op_train_significant = merged_data_double_significant[merged_data_double_significant['Date'] < split_date]
X_ESG_Op_test_significant = merged_data_double_significant[merged_data_double_significant['Date'] >= split_date]

print("Training Period:", X_ESG_Op_train_significant['Date'].min(), "till", X_ESG_Op_train_significant['Date'].max())
print("Testing Period:", X_ESG_Op_test_significant['Date'].min(), "till", X_ESG_Op_test_significant['Date'].max())

# 2. Hgigh Sensitive High Score ESG+Op
dates_high = merged_data_double_high_significant['Date']
if not pd.api.types.is_datetime64_any_dtype(dates_high):
    dates_high = pd.to_datetime(dates_high)

merged_data_double_high_significant['Date'] = pd.to_datetime(
    merged_data_double_high_significant['Date'], 
    format='%Y-%m'  
)

unique_dates_high = np.sort(merged_data_double_high_significant['Date'].unique())

split_idx_high = int(len(unique_dates_high) * 0.8)
split_date_high = unique_dates[split_idx_high]

X_ESG_Op_train_high_significant = merged_data_double_high_significant[merged_data_double_high_significant['Date'] < split_date_high]
X_ESG_Op_test_high_significant = merged_data_double_high_significant[merged_data_double_high_significant['Date'] >= split_date_high]

print("Training Period:", X_ESG_Op_train_high_significant['Date'].min(), "till", X_ESG_Op_train_high_significant['Date'].max())
print("Testing Period:", X_ESG_Op_test_high_significant['Date'].min(), "till", X_ESG_Op_test_high_significant['Date'].max())

# 3. High Sensitive Low Score ESG+Op
dates_low = merged_data_double_low_significant['Date']
if not pd.api.types.is_datetime64_any_dtype(dates_low):
    dates_low = pd.to_datetime(dates_low)

merged_data_double_low_significant['Date'] = pd.to_datetime(
    merged_data_double_low_significant['Date'], 
    format='%Y-%m'
)

unique_dates_low = np.sort(merged_data_double_low_significant['Date'].unique())

split_idx_low = int(len(unique_dates_low) * 0.8)
split_date_low = unique_dates[split_idx_low]

merged_data_double_low_significant.drop(columns=['Return_x'], inplace=True)
merged_data_double_low_significant.rename(columns={'Return_y': 'Return'}, inplace=True)

X_ESG_Op_train_low_significant = merged_data_double_low_significant[merged_data_double_low_significant['Date'] < split_date_low]
X_ESG_Op_test_low_significant = merged_data_double_low_significant[merged_data_double_low_significant['Date'] >= split_date_low]

print("Training Period:", X_ESG_Op_train_low_significant['Date'].min(), "till", X_ESG_Op_train_low_significant['Date'].max())
print("Testing Period:", X_ESG_Op_test_low_significant['Date'].min(), "till", X_ESG_Op_test_low_significant['Date'].max())


# 3.3 Train the two models with OLS and NNs
# for OLS
# 3.3.1 High Sensitive ESG
lags = [1, 3, 12]
results_esg = []
scatter_data_esg = {}  
reg_lines_esg = {}  

df_esg = train_data_esg.reset_index() 
df_esg = df_esg.set_index(['Stock', 'Date'])

for lag in lags:
    df_esg[f'ESG_lag{lag}'] = df_esg.groupby(level='Stock')['ESG'].shift(lag)


for lag in lags:
    col = f'ESG_lag{lag}'
    needed_cols = ['Return', col]

    if col not in df_esg.columns:
        print(f"Column is missiong，skip: {col}")
        continue

    data_lagged = df_esg[needed_cols].dropna()
    scatter_data_esg[lag] = data_lagged  

    if len(data_lagged) < 100:
        print(f"Lag {lag} has not enought sample（n={len(data_lagged)}），skip")
        continue
    
    try:
        data_lagged['const'] = 1  
        model = PanelOLS(
            dependent=data_lagged['Return'],
            exog=data_lagged[[col, 'const']], 
            entity_effects=True
        )
        result = model.fit(
            cov_type="clustered", 
            cluster_entity=True
        )
        
        results_esg.append({
            "Lag": lag,
            "Coef": result.params[col],
            "Pval": result.pvalues[col],
            "Tstat": result.tstats[col],
            "R2": result.rsquared,
            "Nobs": result.nobs,
            "TimePeriod": f"{data_lagged.index.get_level_values('Date').min().date()} - {data_lagged.index.get_level_values('Date').max().date()}"
        })
        
        slope = result.params[col]
        intercept = result.params['const']
        reg_lines_esg[lag] = (slope, intercept)
        
    except Exception as e:
        print(f"Lag {lag} Analysis failed: {str(e)}")
        results_esg.append({
            "Lag": lag,
            "Coef": None,
            "Pval": None,
            "Tstat": None,
            "R2": None,
            "Nobs": len(data_lagged),
            "TimePeriod": None
        })

df_result_esg = pd.DataFrame(results_esg).set_index('Lag')
print("\n Regression Results:")
print(df_result_esg[['Coef', 'Pval', 'R2', 'Nobs', 'TimePeriod']])    

plt.figure(figsize=(10, 6), dpi=300)
colors = {1: 'royalblue', 3: 'coral'}
linestyles = {1: '--', 3: '-'}

for lag in [1, 3]: 
    if lag not in scatter_data_esg or lag not in reg_lines_esg:
        continue
 
    data = scatter_data_esg[lag].reset_index()  
    scatter_data_esg = {k: v for k, v in scatter_data_esg.items() if k in [1, 3]}
    slope, intercept = reg_lines_esg[lag]
    
    xmin = data[f'ESG_lag{lag}'].min()
    xmax = data[f'ESG_lag{lag}'].max()
    xvals = np.linspace(xmin, xmax, 100)
    
    sns.scatterplot(
        x=f'ESG_lag{lag}', y='Return',
        data=data,
        alpha=0.3,
        color=colors[lag],
        label=f'Lag={lag}'
    )

    plt.plot(
        xvals, intercept + slope * xvals,  
        color=colors[lag], 
        lw=3 if lag == 1 else 1.5, 
        alpha=0.8 if lag == 1 else 0.6,  
        linestyle=linestyles[lag],
        label=f'Lag={lag} Fit (β={slope:.4f})'
    )

plt.title('High Sensitive ESG Lagged Effects on Stock Returns', fontsize=14, pad=20)
plt.xlabel('ESG Score (Lagged)', fontsize=10)
plt.ylabel('Return', fontsize=10)
plt.grid(True, alpha=0.3)
plt.legend(frameon=True, facecolor='white', loc='upper left')

plt.savefig(r'/Users/sida/Desktop/ESG/Data/sensitiveG_ESG_Return_Lagged_Effects.png', bbox_inches='tight')
plt.show()


# 3.3.2 High Sensitive High Score ESG PanelOLS
results_high_esg = []
scatter_data_high_esg = {}  
reg_lines_high_esg = {}  

df_high_esg = train_data_high_esg.reset_index() 
df_high_esg = df_high_esg.set_index(['Stock', 'Date'])

for lag in lags:
    df_high_esg[f'ESG_lag{lag}'] = df_high_esg.groupby(level='Stock')['ESG'].shift(lag)

for lag in lags:
    col = f'ESG_lag{lag}'
    needed_cols = ['Return', col]
    
    if col not in df_high_esg.columns:
        print(f"missiong Columns, skip: {col}")
        continue
    
    data_lagged = df_high_esg[needed_cols].dropna()
    scatter_data_high_esg[lag] = data_lagged  
    
    if len(data_lagged) < 100:
        print(f"Lag {lag} has not enought sample（n={len(data_lagged)}），skip")
        continue
    
    try:
        data_lagged['const'] = 1  
        model = PanelOLS(
            dependent=data_lagged['Return'],
            exog=data_lagged[[col, 'const']], 
            entity_effects=True
        )
        result = model.fit(
            cov_type="clustered", 
            cluster_entity=True
        )
        
        results_high_esg.append({
            "Lag": lag,
            "Coef": result.params[col],
            "Pval": result.pvalues[col],
            "Tstat": result.tstats[col],
            "R2": result.rsquared,
            "Nobs": result.nobs,
            "TimePeriod": f"{data_lagged.index.get_level_values('Date').min().date()} - {data_lagged.index.get_level_values('Date').max().date()}"
        })
        
        slope = result.params[col]
        intercept = result.params['const']
        reg_lines_high_esg[lag] = (slope, intercept)
        
    except Exception as e:
        print(f"Lag {lag} Analysis failed: {str(e)}")
        results_high_esg.append({
            "Lag": lag,
            "Coef": None,
            "Pval": None,
            "Tstat": None,
            "R2": None,
            "Nobs": len(data_lagged),
            "TimePeriod": None
        })

df_result_high_esg = pd.DataFrame(results_high_esg).set_index('Lag')
print("\nRegression reults:")
print(df_result_high_esg[['Coef', 'Pval', 'R2', 'Nobs', 'TimePeriod']])

plt.figure(figsize=(10, 6), dpi=300)
colors = {1: 'royalblue', 3: 'coral'}
linestyles = {1: '--', 3: '-'}

for lag in [1, 3]: 
    if lag not in scatter_data_high_esg or lag not in reg_lines_high_esg:
        continue
 
    data = scatter_data_high_esg[lag].reset_index()  
    scatter_data_high_esg = {k: v for k, v in scatter_data_high_esg.items() if k in [1, 3]}
    slope, intercept = reg_lines_high_esg[lag]
    
    xmin = data[f'ESG_lag{lag}'].min()
    xmax = data[f'ESG_lag{lag}'].max()
    xvals = np.linspace(xmin, xmax, 100)
    
    sns.scatterplot(
        x=f'ESG_lag{lag}', y='Return',
        data=data,
        alpha=0.3,
        color=colors[lag],
        label=f'Lag={lag}'
    )

    plt.plot(
        xvals, intercept + slope * xvals,  
        color=colors[lag], 
        lw=3 if lag == 1 else 1.5, 
        alpha=0.8 if lag == 1 else 0.6,  
        linestyle=linestyles[lag],
        label=f'Lag={lag} Fit (β={slope:.4f})'
    )

plt.title('High Sensitive High ESG Lagged Effects on Stock Returns', fontsize=14, pad=20)
plt.xlabel('ESG Score (Lagged)', fontsize=10)
plt.ylabel('Return', fontsize=10)
plt.grid(True, alpha=0.3)
plt.legend(frameon=True, facecolor='white', loc='upper left')

plt.savefig(r'/Users/sida/Desktop/ESG/Data/high_sensitive_high_ESG_Return_Lagged_Effects.png', bbox_inches='tight')
plt.show()


# 3.3.3 High Sensitive Low Score ESG PanelOLS
results_low_esg = []
scatter_data_low_esg = {}  
reg_lines_low_esg = {}  

df_low_esg = train_data_low_esg.reset_index() 
df_low_esg = df_low_esg.set_index(['Stock', 'Date'])

for lag in lags:
    df_low_esg[f'ESG_lag{lag}'] = df_low_esg.groupby(level='Stock')['ESG'].shift(lag)

for lag in lags:
    col = f'ESG_lag{lag}'
    needed_cols = ['Return', col]
    
    if col not in df_low_esg.columns:
        print(f"missing column, skip: {col}")
        continue
    
    data_lagged = df_low_esg[needed_cols].dropna()
    scatter_data_low_esg[lag] = data_lagged  
    
    if len(data_lagged) < 100:
        print(f"Lag {lag} sample not enought（n={len(data_lagged)}），skip")
        continue
    
    try:
        data_lagged['const'] = 1  
        model = PanelOLS(
            dependent=data_lagged['Return'],
            exog=data_lagged[[col, 'const']], 
            entity_effects=True
        )
        result = model.fit(
            cov_type="clustered", 
            cluster_entity=True
        )
        
        results_low_esg.append({
            "Lag": lag,
            "Coef": result.params[col],
            "Pval": result.pvalues[col],
            "Tstat": result.tstats[col],
            "R2": result.rsquared,
            "Nobs": result.nobs,
            "TimePeriod": f"{data_lagged.index.get_level_values('Date').min().date()} - {data_lagged.index.get_level_values('Date').max().date()}"
        })
        
        slope = result.params[col]
        intercept = result.params['const']
        reg_lines_low_esg[lag] = (slope, intercept)
        
    except Exception as e:
        print(f"Lag {lag} Analysis failed: {str(e)}")
        results_low_esg.append({
            "Lag": lag,
            "Coef": None,
            "Pval": None,
            "Tstat": None,
            "R2": None,
            "Nobs": len(data_lagged),
            "TimePeriod": None
        })

df_result_low_esg = pd.DataFrame(results_low_esg).set_index('Lag')
print("\n Regression Results:")
print(df_result_low_esg[['Coef', 'Pval', 'R2', 'Nobs', 'TimePeriod']])

plt.figure(figsize=(10, 6), dpi=300)
colors = {1: 'royalblue', 3: 'coral', 12: 'purple'}
linestyles = {1: '--', 3: '-', 12: '-.'}

for lag in [1, 3, 12]: 
    if lag not in scatter_data_low_esg or lag not in reg_lines_low_esg:
        continue
 
    data = scatter_data_low_esg[lag].reset_index()  
    scatter_data_low_esg = {k: v for k, v in scatter_data_low_esg.items() if k in [1, 3, 12]}
    slope, intercept = reg_lines_low_esg[lag]
    
    xmin = data[f'ESG_lag{lag}'].min()
    xmax = data[f'ESG_lag{lag}'].max()
    xvals = np.linspace(xmin, xmax, 100)
    
    sns.scatterplot(
        x=f'ESG_lag{lag}', y='Return',
        data=data,
        alpha=0.3,
        color=colors[lag],
        label=f'Lag={lag}'
    )
    
    plt.plot(
    xvals, 
    intercept + slope * xvals,  
    color=colors[lag], 
    lw=3 if lag == 1 else 1.5,          
    alpha=0.8 if lag == 1 else 0.6,     
    linestyle=linestyles[lag],          
    label=f'Lag={lag} Fit (β={slope:.4f})'
    )   

plt.title('High Sensitive Low ESG Lagged Effects on Stock Returns', fontsize=14, pad=20)
plt.xlabel('ESG Score (Lagged)', fontsize=10)
plt.ylabel('Return', fontsize=10)
plt.grid(True, alpha=0.3)
plt.legend(frameon=True, facecolor='white', loc='upper left')

plt.savefig(r'/Users/sida/Desktop/ESG/Data/high_sensitive_low_ESG_Return_Lagged_Effects.png', bbox_inches='tight')
plt.show()

### 3.4 double factors model（factor_esg + factor_op）
# for PanelOLS
# 3.4.1 Sensitive ESG + Op PanlOLS
results_double = []

X_ESG_Op_train_clean = X_ESG_Op_train_significant.set_index(
    ['Stock', 'Date'], drop=True  
).sort_index(level=['Stock', 'Date'])

for lag in lags:
    print(f"\n{'='*30} Processing Lag {lag} {'='*30}")
    
    df_lagged = X_ESG_Op_train_clean.copy()
    lag_col = f'ESG_lag{lag}'
    
    df_lagged[lag_col] = df_lagged.groupby('Stock')['ESG'].shift(lag)
    df_lagged[lag_col] = pd.to_numeric(df_lagged[lag_col], errors='coerce') 
    df_clean = df_lagged.dropna(subset=[lag_col, 'smfiv', 'mfis', 'tlm_delta20', 'Return'])
    
    df_clean = df_lagged.dropna(subset=[lag_col, 'smfiv', 'mfis', 'tlm_delta20', 'Return'])
    
    if len(df_clean) == 0:
        print(f"Lag {lag}: Not enough data, skip")
        continue
    
    print(df_clean[[lag_col, 'Return', 'smfiv', 'mfis', 'tlm_delta20']].dtypes)

    print(df_clean[[lag_col, 'Return', 'smfiv', 'mfis', 'tlm_delta20']].dtypes)

    if len(df_clean) == 0:
        print(f"Lag {lag}: not enough data, skip")
        continue

    X_double = df_clean[[lag_col, 'smfiv', 'mfis', 'tlm_delta20']]
    
    vif = pd.DataFrame({
        "Variable": X_double.columns,
        "VIF": [variance_inflation_factor(X_double.values, i) 
                for i in range(X_double.shape[1])]
    })
    
    print(f"\nVIF for Lag {lag}:")
    print(vif)
    
    vif_path = f"/Users/sida/Desktop/ESG/Data/lag_{lag}_double_vif.png"
    
    save_table_as_image(
        vif,
        f"Lag {lag} VIF Results",
        vif_path,
        format_dict={'VIF': '.2e'}  
    )
    try:
        model = PanelOLS(
            dependent=df_clean['Return'],  
            exog=df_clean[[lag_col, 'smfiv', 'mfis', 'tlm_delta20']],
            entity_effects=True  
        ).fit(cov_type="clustered", cluster_entity=True)
    except Exception as e:
        print(f"model failed: {str(e)}")
        continue

    current_result = {
        "Lag": lag,
        "Coef_ESG": model.params.get(lag_col, None),
        "Pval_ESG": model.pvalues.get(lag_col, None),
        "Coef_smfiv": model.params.get('smfiv', None),
        "Coef_mfis": model.params.get('mfis', None),
        "Coef_tlm": model.params.get('tlm_delta20', None),
        "R2": model.rsquared,
        "Nobs": model.nobs
    }
    results_double.append(current_result)

    display_cols = ['Lag', 'Coef_ESG', 'Pval_ESG', 
                    'Coef_smfiv', 'Coef_mfis', 'Coef_tlm',
                    'R2', 'Nobs']
    df_current = pd.DataFrame([current_result])
    
    format_dict = {
        'Coef_ESG': '.4f',
        'Pval_ESG': '.4f',
        'Coef_smfiv': ',.2f',
        'Coef_mfis': ',.2f',
        'Coef_tlm': ',.2f',
        'R2': '.4%',  
        'Nobs': ',d'  
    }
    
    df_result_esg_op = pd.DataFrame(results_double).set_index('Lag').reset_index()
    print("PanelOLS Result for Sensitive ESG + Op:")
    print(df_result_esg_op)

# 3.4.2 Sensitive High ESG + Op PanlOLS
results_double_high = []

X_ESG_Op_train_high_clean = X_ESG_Op_train_high_significant.set_index(
    ['Stock', 'Date'], drop=True  
).sort_index(level=['Stock', 'Date'])

for lag in lags:
    print(f"\n{'='*30} Processing Lag {lag} {'='*30}")
    
    df_lagged = X_ESG_Op_train_high_clean.copy()
    lag_col = f'ESG_lag{lag}'
    
    df_lagged[lag_col] = df_lagged.groupby('Stock')['ESG'].shift(lag)
    df_lagged[lag_col] = pd.to_numeric(df_lagged[lag_col], errors='coerce') 
    df_clean = df_lagged.dropna(subset=[lag_col, 'smfiv', 'mfis', 'tlm_delta20', 'Return'])
    
    if len(df_clean) == 0:
        print(f"Lag {lag}: not enough data, skip")
        continue
    
    print("\n Data Type Check:")
    print(df_clean[[lag_col, 'Return', 'smfiv', 'mfis', 'tlm_delta20']].dtypes)

    print("\n Data Type Check:")
    print(df_clean[[lag_col, 'Return', 'smfiv', 'mfis', 'tlm_delta20']].dtypes)

    if len(df_clean) == 0:
        print(f"Lag {lag}: not enough data, skip")
        continue

    X_double = df_clean[[lag_col, 'smfiv', 'mfis', 'tlm_delta20']]
    
    vif = pd.DataFrame({
        "Variable": X_double.columns,
        "VIF": [variance_inflation_factor(X_double.values, i) 
                for i in range(X_double.shape[1])]
    })
    
    print(f"\nVIF for Lag {lag}:")
    print(vif)
    
    vif_path = f"/Users/sida/Desktop/ESG/Data/lag_{lag}_high_double_vif.png"
    
    save_table_as_image(
        vif,
        f"Lag {lag} VIF检测结果",
        vif_path,
        format_dict={'VIF': '.2e'} 
    )

    try:
        model = PanelOLS(
            dependent=df_clean['Return'],  
            exog=df_clean[[lag_col, 'smfiv', 'mfis', 'tlm_delta20']],
            entity_effects=True  
        ).fit(cov_type="clustered", cluster_entity=True)
    except Exception as e:
        print(f"Modeling Failed: {str(e)}")
        continue

    current_result = {
        "Lag": lag,
        "Coef_ESG": model.params.get(lag_col, None),
        "Pval_ESG": model.pvalues.get(lag_col, None),
        "Coef_smfiv": model.params.get('smfiv', None),
        "Coef_mfis": model.params.get('mfis', None),
        "Coef_tlm": model.params.get('tlm_delta20', None),
        "R2": model.rsquared,
        "Nobs": model.nobs
    }
    results_double_high.append(current_result)

    display_cols = ['Lag', 'Coef_ESG', 'Pval_ESG', 
                   'Coef_smfiv', 'Coef_mfis', 'Coef_tlm',
                   'R2', 'Nobs']
    df_current = pd.DataFrame([current_result])
    
    format_dict = {
        'Coef_ESG': '.4f',
        'Pval_ESG': '.4f',
        'Coef_smfiv': ',.2f',
        'Coef_mfis': ',.2f',
        'Coef_tlm': ',.2f',
        'R2': '.4%',  
        'Nobs': ',d'  
    }

df_result_esg_op_high = pd.DataFrame(results_double_high).set_index('Lag').reset_index()
print("PanelOLS Result for Sensitive High ESG + Op:")
print(df_result_esg_op_high[['Coef_ESG', 'Pval_ESG', 'Coef_smfiv', 'Coef_mfis', 'Coef_tlm', 'R2',
       'Nobs']])

# 3.4.3 Sensitive Low ESG + Op PanlOLS
results_double_low = []

X_ESG_Op_train_low_clean = X_ESG_Op_train_low_significant.set_index(
    ['Stock', 'Date'], drop=True  
).sort_index(level=['Stock', 'Date'])

for lag in lags:
    print(f"\n{'='*30} Processing Lag {lag} {'='*30}")
    
    df_lagged = X_ESG_Op_train_low_clean.copy()
    lag_col = f'ESG_lag{lag}'
    
    df_lagged[lag_col] = df_lagged.groupby('Stock')['ESG'].shift(lag)
    df_lagged[lag_col] = pd.to_numeric(df_lagged[lag_col], errors='coerce') 

    df_clean = df_lagged.dropna(subset=[lag_col, 'smfiv', 'mfis', 'tlm_delta20', 'Return'])
    
    if len(df_clean) == 0:
        print(f"Lag {lag}: Data is insufficient，skip")
        continue
    
    print(df_clean[[lag_col, 'Return', 'smfiv', 'mfis', 'tlm_delta20']].dtypes)

    if len(df_clean) == 0:
        print(f"Lag {lag}: not enough data, skip")
        continue

    X_double = df_clean[[lag_col, 'smfiv', 'mfis', 'tlm_delta20']]
    
    vif = pd.DataFrame({
        "Variable": X_double.columns,
        "VIF": [variance_inflation_factor(X_double.values, i) 
                for i in range(X_double.shape[1])]
    })
    
    print(f"\nVIF for Lag {lag}:")
    print(vif)
    
    vif_path = f"/Users/sida/Desktop/ESG/Data/lag_{lag}_low_double_vif.png"
    
    save_table_as_image(
        vif,
        f"Lag {lag} VIF Test Results", 
        vif_path,
        format_dict={'VIF': '.2e'}  
    )

    try:
        model = PanelOLS(
            dependent=df_clean['Return'],  
            exog=df_clean[[lag_col, 'smfiv', 'mfis', 'tlm_delta20']],
            entity_effects=True  
        ).fit(cov_type="clustered", cluster_entity=True)
    except Exception as e:
        print(f"Modelling failed: {str(e)}")
        continue

    current_result = {
        "Lag": lag,
        "Coef_ESG": model.params.get(lag_col, None),
        "Pval_ESG": model.pvalues.get(lag_col, None),
        "Coef_smfiv": model.params.get('smfiv', None),
        "Coef_mfis": model.params.get('mfis', None),
        "Coef_tlm": model.params.get('tlm_delta20', None),
        "R2": model.rsquared,
        "Nobs": model.nobs
    }
    results_double_low.append(current_result)

    display_cols = ['Lag', 'Coef_ESG', 'Pval_ESG', 
                   'Coef_smfiv', 'Coef_mfis', 'Coef_tlm',
                   'R2', 'Nobs']
    df_current = pd.DataFrame([current_result])
    
    format_dict = {
        'Coef_ESG': '.4f',
        'Pval_ESG': '.4f',
        'Coef_smfiv': ',.2f',
        'Coef_mfis': ',.2f',
        'Coef_tlm': ',.2f',
        'R2': '.4%',  
        'Nobs': ',d' 
    }

df_result_esg_op_low = pd.DataFrame(results_double_low).set_index('Lag').reset_index()

print("PanelOLS Result for Sensitive Low ESG + Op:")
print(df_result_esg_op_low[['Coef_ESG', 'Pval_ESG', 'Coef_smfiv', 'Coef_mfis', 'Coef_tlm', 'R2',
       'Nobs']])

# 3.4.4 Deal with high VIF & significant Granger Causality 
features_esg = merged_data_esg_significant[['ESG']] 

columns_esg = pd.MultiIndex.from_tuples([
    ('ESG', 'ESG_Score')
])

scaler_esg = StandardScaler()
scaled_data_esg = scaler_esg.fit_transform(features_esg)

features_esg_scaled_df = pd.DataFrame(
    scaled_data_esg,
    columns=columns_esg,
    index=merged_data_esg_significant.set_index(['Date', 'Stock']).index  
)

features_esg_op = merged_data_double_significant[['ESG', 'smfiv', 'mfis', 'tlm_delta20']]

columns_esg_op = pd.MultiIndex.from_tuples([
    ('ESG', 'ESG_Score'),          
    ('Option', 'smfiv'),          
    ('Option', 'mfis'),           
    ('Option', 'tlm_delta20')     
])

scaler_esg_op = StandardScaler()
scaled_data_esg_op = scaler_esg_op.fit_transform(features_esg_op)

features_esg_op_scaled_df = pd.DataFrame(
    scaled_data_esg_op,
    columns=columns_esg_op,
    index=merged_data_double_significant.set_index(['Date', 'Stock']).index
)

# PCA Analysis
pca = PCA()

feature_groups_data = [
    ('ESG', features_esg_scaled_df.xs(key='ESG', axis=1, level=0)),  
    ('Option_ESG', features_esg_op_scaled_df.xs(key='Option', axis=1, level=0))
]

for group_name, data_matrix in feature_groups_data:  
    pca.fit(data_matrix)  
    

granger_test = grangercausalitytests(data_matrix[['smfiv', 'mfis']], maxlag=3)
granger_test = grangercausalitytests(data_matrix[['mfis', 'smfiv']], maxlag=3)

granger_test = grangercausalitytests(data_matrix[['mfis', 'tlm_delta20']], maxlag=3)
granger_test = grangercausalitytests(data_matrix[['tlm_delta20', 'mfis']], maxlag=3)

granger_test = grangercausalitytests(data_matrix[['smfiv', 'tlm_delta20']], maxlag=3)
granger_test = grangercausalitytests(data_matrix[['tlm_delta20', 'smfiv']], maxlag=3)


explained_variance = pca.explained_variance_ratio_
cumulative_variance = explained_variance.cumsum()

print("Explained variance ratio for each component:", explained_variance)
print("Cumulative explained variance:", cumulative_variance)

plt.figure(figsize=(8, 6))
plt.plot(range(1, len(cumulative_variance) + 1), cumulative_variance, marker='o', linestyle='--', color='#006400')
plt.title('Cumulative Explained Variance by PCA Components')
plt.xlabel('Number of Components')
plt.ylabel('Cumulative Explained Variance')
plt.grid(True)
plt.savefig("/Users/sida/Desktop/ESG/Data/pca_variance_plot.png")
plt.show()

# After

New_features = [('ESG', 'ESG_Score'), ('Option', 'smfiv')]
features_esg_smfiv_scaled_df = features_esg_op_scaled_df.loc[:, New_features]
features_esg_smfiv_scaled_df.columns = features_esg_smfiv_scaled_df.columns.get_level_values(1)

pca.fit(features_esg_smfiv_scaled_df)


explained_variance = pca.explained_variance_ratio_
cumulative_variance = explained_variance.cumsum()

print("Explained variance ratio for each component:", explained_variance)
print("Cumulative explained variance:", cumulative_variance)

plt.figure(figsize=(8, 6))
plt.plot(range(1, len(cumulative_variance) + 1), cumulative_variance, marker='o', linestyle='--', color='#006400')
plt.title('Cumulative Explained Variance by ESG and smfiv')
plt.xlabel('Number of Components')
plt.ylabel('Cumulative Explained Variance')
plt.grid(True)
plt.savefig("/Users/sida/Desktop/ESG/Data/drop2_plot.png")
plt.show()

# 3.4.6 for Significant ESG + smfiv
results_esg_smfiv = []
scatter_data_esg_smfiv = {}  
reg_lines_esg_smfiv = {}  

df_esg_smfiv= X_ESG_Op_train_clean.reset_index() 
df_esg_smfiv = df_esg_smfiv.set_index(['Stock', 'Date'])

df_high_esg_smfiv= X_ESG_Op_train_high_clean.reset_index() 
df_high_esg_smfiv = df_high_esg_smfiv.set_index(['Stock', 'Date'])

df_low_esg_smfiv= X_ESG_Op_train_low_clean.reset_index() 
df_low_esg_smfiv = df_low_esg_smfiv.set_index(['Stock', 'Date'])

analysis_config = {
    'ESG_smfiv': {
        'df': df_esg_smfiv[['ESG', 'smfiv', 'Return']],
        'colors': {1: 'steelblue', 3: 'orange', 12: 'mediumpurple'},
        'linestyles': {1: '--', 3: '-', 12: '-.'}
    },
    'High_ESG_smfiv': {
        'df': df_high_esg_smfiv[['ESG', 'smfiv', 'Return']],
        'colors': {1: 'darkgreen', 3: 'goldenrod', 12: 'firebrick'},
        'linestyles': {1: '--', 3: '-', 12: '-.'}
    },
    'Low_ESG_smfiv': {
        'df': df_low_esg_smfiv[['ESG', 'smfiv', 'Return']],
        'colors': {1: 'royalblue', 3: 'coral', 12: 'purple'},
        'linestyles': {1: '--', 3: '-', 12: '-.'}
    }
}


final_results = {}
lags = [1, 3, 12]
save_base_path = '/Users/sida/Desktop/ESG/Data'

for group_name, config in analysis_config.items():
    print(f"\n{'='*30} analyses {group_name} {'='*30}")
    try:
        results, scatter, lines, vif_tables = analyze_esg_lag_effects(
            df_group=config['df'],
            group_name=group_name,
            lags=lags,
            colors=config['colors'],
            linestyles=config['linestyles'],
            save_path=save_base_path
        )

        final_results[group_name] = {
            'results': results,
            'scatter_data': scatter,
            'reg_lines': lines,
            'vif_tables': vif_tables
        }
        print(f"Finish Analysis: {group_name}")
    except Exception as e:
        print(f"failures in Analysis: {group_name} - mistakes: {e}")

print("Final Results:")
for group_name, group_results in final_results.items():
    print(f"\n{'='*40}")
    print(f"groups: {group_name}")
    print(f"{'='*40}")

    if 'results' in group_results:
        regression_summary = []
        for lag, result in group_results['results'].items():
            try:

                if hasattr(result, 'rsquared') and hasattr(result, 'pvalues'):
                    summary = {
                        'Lag': lag,
                        'R-squared': result.rsquared,
                        'Adj R-squared': result.rsquared_adj if hasattr(result, 'rsquared_adj') else 'N/A',
                        'P-value': result.pvalues.min() if not result.pvalues.empty else 'N/A',
                        'F-statistic': result.fvalue if hasattr(result, 'fvalue') else 'N/A'
                    }
                    regression_summary.append(summary)
                else:
                    print(f"warnings, can not withdraw (Lag {lag})，try to print")

            except Exception as e:
                print(f"Mistaks: (Lag {lag}) - {e}")

    
        if regression_summary:
            df_regression = pd.DataFrame(regression_summary)
            print("Regression Tables:")
            print(df_regression.to_string(index=False))
        else:
            print("Invalid Regression Results")
    else:
        print("Invalid Regression Results")

    try:
        if 'reg_lines' in group_results:
            plt.figure(figsize=(8, 6))
            for key, line in group_results['reg_lines'].items():

                if isinstance(line, tuple):
                    x_values, y_values = line
                elif isinstance(line, dict) and 'x' in line and 'y' in line:
                    x_values, y_values = line['x'], line['y']
                else:
                    print(f"Warnings: unknown Regression, skip {key}")
                    continue
                
                plt.plot(x_values, y_values, label=f"Lag {key}", linestyle='-', marker='o')

            plt.title(f"Regression - {group_name}")
            plt.xlabel("X (Factor)")
            plt.ylabel("Y (Return)")
            plt.legend()
            plt.grid(True)
            plot_path = f"{save_base_path}/Regression_Results_{group_name}.png"
            plt.savefig(plot_path)
            plt.show()
            print(f"Regression PIctures saved: {plot_path}")
        else:
            print("invalid regression data")
    except Exception as e:
        print(f"Mistakes when sourcing: {e}")


    try:
        if 'vif_tables' in group_results:
            vif_summary = []
            for lag, vif_table in group_results['vif_tables'].items():
                vif_df = pd.DataFrame(vif_table)
                vif_df['Lag'] = lag
                vif_summary.append(vif_df)

            if vif_summary:
                df_vif = pd.concat(vif_summary, ignore_index=True)
                print("VIF comprehensive table:")
                print(df_vif.to_string(index=False))
            else:
                print("invalid VIF data")
        else:
            print("No VIF Table")
    except Exception as e:
        print(f"Mistakes when sourcing VIF table: {e}")

print("Finished Building Models")
        

# Step 4: Portfolio Optimization

data_sources = [

    {'prefix': 'esg', 'features': ['ESG'], 'double': False},
    {'prefix': 'esg_smfiv', 'features': ['ESG', 'smfiv'], 'double': True}
]


data_dict = {}
for source in data_sources:
    for category in ['low', 'high', '']:
        
        name = f"{category}_{source['prefix']}".strip('_')

        if source['double']:
            if category == '':
                df_name = f"merged_data_double_significant"
            else:
                df_name = f"merged_data_double_{category}_significant"
                
        else:
            if category == '':
                df_name = f"merged_data_esg_significant"
            else:
                df_name = f"merged_data_{category}_esg_significant"

        data_dict[name] = globals().get(df_name)

print("Data Source:", {key: type(value).__name__ for key, value in data_dict.items()})

# Step 4.0: Define the configurations with adjusted feature handling
betas_nm = (0.9, 0.999) 
model_configs = {
    'PanelOLS': {
        'model_type': 'PanelOLS',
        'entity_effects': False,  
        'time_effects': False,    
        'cov_type': 'clustered',     
        'regularization': {
            'method': 'betas_nm', 
            'params': betas_nm
        }
    },
    
    'NeuralNetwork': {
        'model_type': 'NeuralNetwork',
        'network_architecture': {
            'hidden_layers': [64, 32],  
            'activation': 'swish',           
            'dropout_rate': 0.2,
            'custom_loss': True  
        },
        'training_params': {
            'epochs': 1,
            'batch_size': 32,
            'learning_rate': 0.001,
            'early_stopping': True,
            'use_gpu': True  
        },
        'regularization': {
            'method': 'l2',
            'lambda': 0.01
        }
    }
}

BASE_CONFIG = {

    'preprocessing': {
        'winsorize': True,
        'winsorize_quantile': 0.01,
        'scaling_method': 'robust'
    },
    
    'feature_generation_rules': {
        'interaction_pairs': [('ESG', 'smfiv')]  
    },
    
    'optimization': {
        'max_weight': 0.05,
        'turnover_limit': 0.2
    }
}

feature_engineering_configs = {
    'ESG_only': {
        'dataset': 'esg',
        'base_features': ['ESG'],
        'preprocessing': BASE_CONFIG['preprocessing'],
        'feature_generation': {'interactions': []}
    },
    'High_ESG_only': {
        'dataset': 'high_esg',
        'base_features': ['ESG'],
        'preprocessing': BASE_CONFIG['preprocessing'],
        'feature_generation': {'interactions': []}
    },
    'Low_ESG_only': {
        'dataset': 'low_esg',
        'base_features': ['ESG'],
        'preprocessing': BASE_CONFIG['preprocessing'],
        'feature_generation': {'interactions': []}
    },
    
    'esg_smfiv_adv': {
        **BASE_CONFIG,
        'dataset': 'esg_smfiv',
        'base_features': ['ESG', 'smfiv'],
        'feature_generation': {'interactions': [('ESG', 'smfiv')]}
    },
    'High_ESG_smfiv_adv': {
        **BASE_CONFIG,
        'dataset': 'high_esg_smfiv',
        'base_features': ['ESG', 'smfiv'],
        'feature_generation': {'interactions': [('ESG', 'smfiv')]}
    },
    'Low_ESG_smfiv_adv': {
        **BASE_CONFIG,
        'dataset': 'low_esg_smfiv',
        'base_features': ['ESG', 'smfiv'],
        'feature_generation': {'interactions': [('ESG', 'smfiv')]}
    }
}

unwanted_keys = ['model_type', 'cov_type', 'regularization']
configurations = {}
for feat_name, feat_config in feature_engineering_configs.items():
    for model_name, model_config in model_configs.items():
        config_key = f"{model_name}|{feat_name}"
        print(f'config_key', config_key)

        mc = copy.deepcopy(model_config)

        fixed_effects = mc.pop('fixed_effects', {})  

        mc = {**mc, **fixed_effects}
        model_params_filtered = {k: v for k, v in mc.items() if k not in unwanted_keys}
        configurations[config_key] = {
            'alpha_estimation_method': model_config['model_type'],
            'pred_horizon': 1,
            'feature_names': feat_config['base_features'],  
            'dataset': feat_config['dataset'],
            'level_winsorize': BASE_CONFIG['preprocessing']['winsorize_quantile'],
            'look_back_prm': 252,
            'n_stocks': 155,
            'optim_params': {'betas_nm': betas_nm},
            'rolling_params': {'window_size': 12, 'step_size': 1},
            'optimization': BASE_CONFIG['optimization'],
            'model_params': model_params_filtered,   
            'feature_generation': feat_config.get('feature_generation', {}),
            'base_config': BASE_CONFIG
        }
print("Generated configurations:")

# Step 4.1: Test significance for all models
master_stats = []

for key, df in data_dict.items():

    df_renamed = df.copy()
    df_renamed.rename(columns={'Return':'RET'}, inplace=True)
    df_renamed['Date'] = pd.to_datetime(df_renamed['Date'])
    df_renamed.set_index(['Date','Stock'], inplace=True)
    df_renamed = df_renamed.sort_index(level='Date')

    for config_name, config in configurations.items():
        if config['dataset'] != key:
            continue

        features = config['feature_names']
        exog = (
            df_renamed[features]
            .reset_index()
            .set_index(['Stock','Date'])
        )
        dep = (
            df_renamed['RET']
            .reset_index()
            .set_index(['Stock','Date'])
        )
        
        X = exog.values
        y = dep.values
        
        X_train, X_val, y_train, y_val = train_test_split(
            X, y,
            test_size=0.2,
            shuffle=False
        )
        if config['alpha_estimation_method'] == 'NeuralNetwork':

            with tf.device('/GPU:0' if config['model_params']['training_params']['use_gpu'] else '/CPU:0'):
                nn = NeuralNetwork(dep, exog, **config['model_params'])
                nn.fit()  # 训练过程里会自己划分 train/val
                history = nn.history
 
            train_mse = history.history['loss'][-1]
            val_mse   = history.history.get('val_loss', [None])[-1]

            y_train_pred = nn.model.predict(X_train, verbose=0)  
            y_val_pred = nn.model.predict(X_val, verbose=0)
            
            train_r2 = r2_score(y_train, y_train_pred)
            val_r2 = r2_score(y_val, y_val_pred) if val_mse is not None else None  
            
            master_stats.append({
                'dataset'      : key,
                'config_name'  : config_name,
                'model_type'   : 'NeuralNetwork',
                'train_mse'    : train_mse,
                'val_mse'      : val_mse,
                'train_r2'     : train_r2,
                'val_r2'       : val_r2,  
                'epochs'       : len(history.history['loss'])
            })

        else:

            panel = PanelOLS(
                dep, exog,
                entity_effects=False,
                time_effects=False
            )
            res = panel.fit(cov_type='clustered')

            master_stats.append({
                'dataset'      : key,
                'config_name'  : config_name,
                'model_type'   : 'PanelOLS',
                'f_pvalue'     : res.f_statistic.pval,
                'r2_overall'   : res.rsquared,
                'r2_within'    : res.rsquared_within,
                'r2_between'   : res.rsquared_between,
                'coeffs'       : res.params.to_dict(),
                'coeff_pvals'  : res.pvalues.to_dict(),
            })


stats_df = pd.DataFrame(master_stats)
stats_df.to_csv('full_sample_model_stats.csv', index=False)
print(stats_df)


stats_df = pd.read_csv('full_sample_model_stats.csv')


ols_df = stats_df[stats_df['model_type'] == 'PanelOLS'].copy()
nn_df  = stats_df[stats_df['model_type'] == 'NeuralNetwork'].copy()


with pd.ExcelWriter('model_stats.xlsx', engine='openpyxl') as writer:
    ols_df.to_excel(writer, sheet_name='PanelOLS_Results', index=False)
    nn_df .to_excel(writer, sheet_name='NeuralNetwork_Results', index=False)

    workbook = writer.book
    for sheet_name, df in [('PanelOLS_Results', ols_df), ('NeuralNetwork_Results', nn_df)]:
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        dims = {}
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                text = str(cell) if cell is not None else ""
                dims[i] = max(dims.get(i, 0), len(text))
        for i, width in dims.items():
            ws.column_dimensions[get_column_letter(i+1)].width = width + 2


# 4.2 Predict Returns
factors_esg = factors_esg.sort_index()
non_zero_factors_esg = factors_esg[(factors_esg != 0).any(axis=1)]
zero_factors_esg = factors_esg[(factors_esg == 0).all(axis=1)]
factors_esg_combined = pd.concat([non_zero_factors_esg, zero_factors_esg])
factors_esg_unique = factors_esg_combined[~factors_esg_combined.index.duplicated(keep='first')]
factors_esg_unique = factors_esg_unique.sort_index()
complete_index = pd.date_range(start=factors_esg_unique.index.min(), end=end_date, freq='ME').strftime('%Y-%m')
factors_esg_unique = factors_esg_unique.reindex(complete_index, fill_value=0)
dates = pd.to_datetime(returns.index)      
periods = dates.to_period('M')              
timestamps = periods.to_timestamp()         
window_size = 12  
step_size = 1
start = time.time()
optimized_weights_list = []
window_weights = []
optimized_weights_dict = {}
all_optimized_portfolios = {}
all_non_zero_weight_tickers = []
all_optimized_weights = {}
expanded_y_train_sr = []
all_y_pred = {}
mse_results = {}
metrics_df = pd.DataFrame(columns=['config_name', 'r_squared', 'mse'])
targets = Stoxx600[['Date', 'Sharpe Ratio']]
targets = targets[1:]
targets.set_index(['Date'], inplace=True)
master_results = [] 
optimization_results = {}
new_returns = {
    config_name: returns.copy()
    for config_name, config in configurations.items()
}

# Model and Predict Returns
for key, df in data_dict.items():
    df_renamed = df.copy()

    df_renamed.rename(columns={'Return': 'RET'}, inplace=True)
    df_renamed['Date'] = pd.to_datetime(df_renamed['Date'], errors='coerce')
    df_renamed.set_index(['Date', 'Stock'], inplace=True)
    df_renamed = df_renamed.sort_index(level='Date')
         
    for i in tqdm(range(0, len(dates) - window_size, step_size)):

        n_windows = len(dates) - window_size

        print_memory_usage()
        y_train_list = []
        current_dates = dates[i:i + window_size]
        start_date = current_dates.min()
        print(f'start_date', start_date)
        end_date = current_dates.max()
        print(f'end_date', end_date)
        window_data = df_renamed.loc[pd.IndexSlice[start_date:end_date, :]]
        print(f'window_data', window_data)
        window_data = window_data.swaplevel('Stock', 'Date').sort_index()

        
        for config_name, config in configurations.items():
            
            if config['dataset'] != key:
               continue 
            features = config['feature_names']
            X_train = pd.DataFrame(columns=features)
            required_features = config['feature_names']
            missing_features = [f for f in required_features if f not in df_renamed.columns]
            if missing_features:
                print(f"skip {config_name}，lacking features: {missing_features}")
                continue  
            print("Processing dataset:", key)
            print("Using configuration:", config['dataset'])
            print("Using configuration的key:", config_name)
            print("Expected features:", config['feature_names'])

            required_features = config['feature_names'] 
            missing_features = [f for f in required_features if f not in window_data.columns]
            if missing_features:
                print(f"skip {config_name}，lacking features: {missing_features}")
                continue

            print(X_train.shape) 
            print(X_train.columns)

            next_date = end_date + pd.DateOffset(months=1)
            next_period = next_date.to_period('M').to_timestamp() 
            
            exog_original = window_data[features].copy()
            exog = exog_original.groupby(level='Stock').mean() 
            exog = (
                exog_original
                  .groupby(level='Stock').mean()
                  .assign(Date=next_period)
                  .reset_index()
                  .set_index(['Stock','Date'])
            )

            dependent_original = window_data['RET'].copy()
            dependent_original = dependent_original.groupby(level='Stock').mean() 
            mi = pd.MultiIndex.from_product(
                [dependent_original.index, [next_period]],
                names=['Stock','Date']
            )
            dependent = pd.Series(
                data= dependent_original.values,
                index=mi,
                name=dependent_original.name
            )

            if config['alpha_estimation_method'] == 'NeuralNetwork':
                print("\n=== NeuralNetwork Modelling Starts ===")
                if config['model_params']['training_params']['use_gpu']:
                    with tf.device('/GPU:0'):
                        model = NeuralNetwork(dependent, exog, **config['model_params'])
                        model.fit()
            else:
                print("\n=== PanelOLS Modelling Starts ===")

                model = PanelOLS(
                    dependent=dependent,
                    exog=exog,  
                    entity_effects=False,
                    time_effects=False
                ).fit(cov_type='clustered')   

           
            y_pred = model.predict(exog=exog)  
            y_pred_df = pd.DataFrame(y_pred, index=exog.index, columns=['prediction'])
            pred_wide = y_pred_df.unstack(level='Stock')
            pred_wide.index = [ next_period ]
            pred_wide.columns = pred_wide.columns.droplevel(0)
            
            period_str = next_period.strftime('%Y-%m')

            orig_window = returns.loc[period_str].copy()
            returns_pred = orig_window.copy().to_frame()
            returns_pred = returns_pred.T
            returns_pred.update(pred_wide)
            print(config['alpha_estimation_method'], f'Modelling Finished')
            
            new_returns[config_name].loc[
                returns_pred.index, returns_pred.columns
            ] = returns_pred.values

# 4.3 Optimize Portfolio    
for config_name, df_pred in new_returns.items():
    optimized_weights_df = pd.DataFrame(index=df_pred.index, columns=df_pred.columns, dtype=float)    
   
    cov_matrix = df_pred.cov().values
    mean_returns = df_pred.mean().values

    objective = lambda w: -(
         (df_pred.values.dot(w) - risk_free_rate['1mth Bund']/12).mean() /
         np.sqrt(w.dot(cov_matrix).dot(w))
    )

    for date in df_pred.index:  
        x0 = initial_weights.loc[date].reindex(df_pred.columns, fill_value=0).values
        constraints = {'type': 'eq','fun': lambda w: np.sum(w) - 1.0}
        
        bounds = [
        (0.0, 0.0) if w0 == 0 else (-1.0, 1.0)
        for w0 in x0
        ]
        
        optimized_result = minimize(
            fun=objective,
            x0=x0,
            method='SLSQP',
            bounds = bounds,
            constraints=constraints
        )

        optimized_weights = optimized_result.x
        optimized_weights_df.loc[date] = optimized_weights

# Step 5: Portfolio Performance Analysis
    optimized_portfolio_return = (df_pred * optimized_weights_df).sum(axis=1)
    optimized_portfolio_return = optimized_portfolio_return.replace([np.inf, -np.inf], 0)
    
    all_optimized_portfolios[config_name] = {'{config_name} Optimized Portfolio': optimized_portfolio_return}
    cumulative_performance = calculate_cumulative_performance(optimized_portfolio_return)
    max_drawdown = calculate_max_drawdown(cumulative_performance)
    cumulative_performance = cumulative_performance[-1]
    sharpe_ratio = calculate_sharpe_ratio(optimized_portfolio_return, risk_free_rate['1mth Bund'])
    
   
    
    optimization_results[config_name] = {
        'Cumulative Performance': cumulative_performance,
        'Sharpe Ratio': sharpe_ratio,
        'Max Drawdown': max_drawdown
    }
   
    print(f"Results for {config_name}:")
    print(f"Cumulative Performance':", cumulative_performance)
    print(f"Sharpe Ratio:", sharpe_ratio)
    print(f"Max Drawdown:", max_drawdown)
   
     
# original portfolio performance analysis
original_cumulative_performance = calculate_cumulative_performance(original_portfolio_return)
original_max_drawdown = calculate_max_drawdown(original_cumulative_performance)
original_cumulative_performance = original_cumulative_performance[-1]
original_sharpe_ratio = calculate_sharpe_ratio(original_portfolio_return, risk_free_rate['1mth Bund'])


optimization_results['Original Portforlio'] = {
    'Cumulative Performance': original_cumulative_performance,
    'Sharpe Ratio': original_sharpe_ratio,
    'Max Drawdown': original_max_drawdown
}

print(f"Results for Original Portfolio:")
print(f"Cumulative Performance':", original_cumulative_performance)
print(f"Sharpe Ratio:", original_sharpe_ratio)
print(f"Max Drawdown:", original_max_drawdown)

